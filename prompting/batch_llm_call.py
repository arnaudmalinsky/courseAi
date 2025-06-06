import asyncio
import json
from tqdm import tqdm
import time
from pathlib import Path
import openpyxl
from datetime import datetime
import logging

import openai
from openai import OpenAI, OpenAIError
import pandas as pd
from langchain_openai import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.schema import OutputParserException
from langchain.output_parsers import PydanticOutputParser
from pydantic import BaseModel, ValidationError
from langchain.chains import LLMChain

from .law_reference_identification_prompt import (
    LAW_REF_PROMPT,LAW_REF_HEADER_FORMAT
)
from .sumup_prompt import (
    SUMUP_PROMPT,
    SUMUP_HEADER_FORMAT
)



class ExcelManager():
    def __init__(self, output_excel_file_path, header_format):
        self.header_format=header_format
        self.output_excel_file_path = self.get_output_path(output_excel_file_path)
        self.flag_exists = self.get_flag_exists()
        self.workbook, self.worksheet=self._get_excel_workbook()
        self.beginning_index=self._get_beginning_index()
        self.inspect_or_create_header()
        self.new_excel_path = self._get_timestamp()

    def get_output_path(self, output_excel_file_path):
        try:
            return Path(output_excel_file_path)
        except TypeError as e:
            logging.warning(f"No output path was given, error: {e}")
            return None

    def get_flag_exists(self):
        if self.output_excel_file_path:
            flag_exists = self.output_excel_file_path.is_file()
            logging.warning(f"File is found bool: {flag_exists}")
        else:
            logging.warning(f"No output excel path given")
            flag_exists=False
        return flag_exists
    
    def _get_timestamp(self):
        now = datetime.now()
        ts = datetime.timestamp(now)
        new_excel_path=(
            f"../courseai_data/llm_result_all_corpus_{str(ts)}.xlsx"
        )
        return new_excel_path

    def _get_beginning_index(self):
        logging.warning(f"max row of existing file is {self.worksheet.max_row}")
        beginning_index = self.worksheet.max_row - 1
        logging.warning(f"beginning index should be {beginning_index}")
        return beginning_index

    def _get_excel_workbook(self):
        if self.flag_exists is True:
            wb = openpyxl.load_workbook(self.output_excel_file_path)
            logging.warning(f"Existing excel is loaded")
        else:
            wb = openpyxl.Workbook()
            logging.warning(f"New workbook is created")
        return wb, wb.active
    
    def inspect_or_create_header(self):
        values_list = list(self.worksheet.values)
        self.flag_header = True
        if len(values_list)==0:
            self.flag_header = False
        elif list(values_list[0]) != self.header_format: #[:len(self.header_format)]

            self.flag_header = False

        if self.flag_header is False:
            logging.warning(f"File format can't be used, new workbook is created")
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.append(self.header_format)
            self.beginning_index=0
        else:
            logging.warning("Header of excel is: %s",  values_list[0])
            logging.warning("Last line values are: %s",  values_list[-1])
            logging.warning(f"File format is correct, results will be stacked")
    
    def batch_append(self, batch_results_df):
        # for idx, row in batch_results_df.iterrows():
        #     self.worksheet.append(
        #         row.to_list()
        #     )
        for idx, row in batch_results_df.iterrows():
            row_data = [row[col] for col in self.header_format]  # Only take the columns in the expected header order
            self.worksheet.append(row_data)
        self.workbook.save(self.new_excel_path)


class LawRefResponseSchema(BaseModel):
    flag_law: bool
    label: str
    law_description:str
    corpus: str
    institution: str
    law_type: str
    location: str
    date: str

class SumUpResponseSchema(BaseModel):
    text_sumup: str

def get_chain(instruction, llm, parser):
    

    prompt = PromptTemplate(
        template=(
            instruction + 
            "\n Instructions de format : {format_instructions}"+
            "\n Le texte à analyser : {query}\n"
        ),
        input_variables=["query"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )

    chain = prompt | llm | parser
    # chain = LLMChain(llm=llm, prompt=prompt)
    return chain

def get_sumup_chain(instruction, llm):
    prompt = PromptTemplate(
        template=(instruction + "{query}"),
        input_variables=["query"],
    )

    chain = prompt | llm
    # chain = LLMChain(llm=llm, prompt=prompt)
    return chain


async def run_chain(chain, text, retries=2, delay=2):
    if len(text) < 100:
        logging.info("Skipping short text (less than 100 characters).")
        return None
    for attempt in range(retries):
        try:
            result = await chain.ainvoke({'query': text})
            return result
        except OpenAIError as e:
            logging.warning(f"Rate limit hit for try {attempt}. Retrying in {delay} seconds... Error : {e}")
            await asyncio.sleep(delay)
        except (ValidationError, OutputParserException) as e:
            logging.warning(f"Error processing \n{e}")
            return None
    logging.warning("Max retries reached. Skipping.")
    return None

def format_and_merge_results(df, batch_results, idx_batch_beginning, parser_class):
    default_dict = {field: None for field in parser_class.model_fields}
    result_dict_list = []
    idx = idx_batch_beginning
    for result in batch_results:
        if result is None:
            default_dict["unique_id"]=idx
            result_dict_list.append(
                default_dict
            )
        else:
            result_dict=result.dict()
            result_dict['unique_id']=idx
            result_dict_list.append(result_dict)
        idx+=1
        
    result_df = pd.DataFrame(result_dict_list)
    input_and_result = df.merge(result_df, on="unique_id", how='inner')
    return input_and_result

# Define a function to run multiple chains concurrently
async def run_multiple_chains(chain,texts):
    tasks = [run_chain(chain, text) for text in texts]
    results = await asyncio.gather(*tasks)
    return results

def dummy_test(texts):
    return [None for text in texts]

def get_results(
        df, 
        chain, 
        batches, 
        excel_output_object, 
        batch_size,
        parser_class
    ):

    all_results = []
    idx=excel_output_object.beginning_index
    for batch in tqdm(batches):
        batch_results = asyncio.run(run_multiple_chains(chain, batch))
        all_results.extend(batch_results)
        batch_results_df=format_and_merge_results(df, batch_results,idx,parser_class)
        excel_output_object.batch_append(batch_results_df )
        # time.sleep(2)
        idx+=batch_size
    return all_results

def get_batches(df, batch_size, text_column="Text"):
    return [df[text_column][i:i + batch_size].fillna("").tolist() for i in range(0, len(df), batch_size)]

def get_formatted_data(input_excel_file_path,excel_output_object, limit):
    sample_df = pd.read_excel(input_excel_file_path)
    
    sample_df.reset_index(inplace=True)
    sample_df.rename(columns={'index':'unique_id'}, inplace=True)
    
    sample_df=sample_df[excel_output_object.beginning_index:]
    
    if limit:
        sample_df=sample_df[:limit]

    return sample_df


# Function to process a DataFrame using the LLM
def batch_call(
        open_ai_key,
        input_excel_file_path,
        output_excel_file_path, 
        flag_law_ref,
        batch_size = 500, 
        limit=None
    ):
    
    
    llm = ChatOpenAI(
        model_name="gpt-4o",#-mini
        openai_api_key = open_ai_key,
        temperature = 0
    )

    # Ensure the DataFrame has a 'text' column
    

    if flag_law_ref:
        excel_output_object = ExcelManager(output_excel_file_path, LAW_REF_HEADER_FORMAT)
        df = get_formatted_data(input_excel_file_path, excel_output_object,limit)
        if 'Text' not in df.columns:
            raise ValueError("DataFrame must contain a 'text' column")
        instruction =LAW_REF_PROMPT
        batches_list = get_batches(df, batch_size, 'Text')
        parser = PydanticOutputParser(pydantic_object=LawRefResponseSchema)
        chain=get_chain(instruction, llm, parser)
        results_list = get_results(
            df,
            chain, 
            batches_list, 
            excel_output_object, 
            batch_size,
            LawRefResponseSchema
        )
    else:
        excel_output_object = ExcelManager(output_excel_file_path, SUMUP_HEADER_FORMAT)
        df = get_formatted_data(input_excel_file_path, excel_output_object,limit)
        if 'concat_texts' not in df.columns:
            raise ValueError("DataFrame must contain a 'concat_texts' column")
        instruction =SUMUP_PROMPT
        batches_list = get_batches(df, batch_size, 'concat_texts')
        parser = PydanticOutputParser(pydantic_object=SumUpResponseSchema)
        chain=get_chain(instruction, llm, parser)

        results_list = get_results(
            df,
            chain, 
            batches_list, 
            excel_output_object, 
            batch_size,
            SumUpResponseSchema
        )

    return results_list